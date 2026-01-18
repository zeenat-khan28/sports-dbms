"""Export API endpoints for downloading data."""
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import Optional
from datetime import datetime
import io

from app.db.mongodb import get_submissions_collection, get_participation_collection
from app.db.postgres import get_postgres_session
from app.models.sql_models import User, Event
from app.core.security import get_current_admin_user

router = APIRouter(prefix="/export", tags=["Export"])


@router.get("/submissions/csv")
async def export_submissions_csv(
    branch: Optional[str] = Query(None, description="Filter by branch"),
    status: str = Query("approved", description="Filter by status"),
    current_user: User = Depends(get_current_admin_user)
):
    """Export submissions as CSV (Admin only)."""
    collection = get_submissions_collection()
    
    query = {"status": status}
    if branch:
        query["branch"] = branch
    
    cursor = collection.find(query).sort("sln", 1)
    submissions = await cursor.to_list(length=1000)
    
    output = io.StringIO()
    
    # College Header
    output.write("RV COLLEGE OF ENGINEERING (RVCE) - SPORTS DEPARTMENT\n")
    output.write(f"Student Registration List ({status.upper()}) - Generated: {datetime.now().strftime('%d/%m/%Y')}\n")
    output.write("\n")
    
    # Headers
    headers = [
        "SLN", "Name", "USN", "Branch", "Semester",
        "DOB", "Blood Group", "Phone", "Parent Name", "Mother Name"
    ]
    output.write(",".join(headers) + "\n")
    
    for sub in submissions:
        row = [
            str(sub.get("sln", "")),
            sub.get("student_name", ""),
            sub.get("usn", ""),
            sub.get("branch", ""),
            str(sub.get("semester", "")),
            sub.get("date_of_birth", ""),
            sub.get("blood_group", ""),
            sub.get("phone", ""),
            sub.get("parent_name", ""),
            sub.get("mother_name", "")
        ]
        output.write(",".join([f'"{cell}"' for cell in row]) + "\n")
    
    output.seek(0)
    filename = f"students_{status}_{datetime.now().strftime('%Y%m%d')}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/submissions/excel")
async def export_submissions_excel(
    branch: Optional[str] = Query(None, description="Filter by branch"),
    status: str = Query("approved", description="Filter by status"),
    include_header: bool = Query(True, description="Include college header"),
    include_footer: bool = Query(True, description="Include footer"),
    current_user: User = Depends(get_current_admin_user)
):
    """Export submissions as Excel with optional college header/footer (Admin only)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")
    
    collection = get_submissions_collection()
    
    query = {"status": status}
    if branch:
        query["branch"] = branch
    
    cursor = collection.find(query).sort("sln", 1)
    submissions = await cursor.to_list(length=1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    title_font = Font(bold=True, size=16, color="1a1a2e")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row_num = 1
    
    # College Header
    if include_header:
        ws.merge_cells('A1:J1')
        ws['A1'] = "RV COLLEGE OF ENGINEERING (RVCE)"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:J2')
        ws['A2'] = "SPORTS DEPARTMENT - STUDENT REGISTRATION"
        ws['A2'].font = Font(bold=True, size=14, color="3366ff")
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A3:J3')
        ws['A3'] = f"Report Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
        ws['A3'].alignment = Alignment(horizontal='center')
        
        row_num = 5
    
    # Headers
    headers = ["SLN", "Name", "USN", "Branch", "Semester", "DOB", "Blood Group", "Phone", "Parent", "Mother"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row_num += 1
    
    # Data rows
    for sub in submissions:
        data = [
            sub.get("sln", ""),
            sub.get("student_name", ""),
            sub.get("usn", ""),
            sub.get("branch", ""),
            sub.get("semester", ""),
            sub.get("date_of_birth", ""),
            sub.get("blood_group", ""),
            sub.get("phone", ""),
            sub.get("parent_name", ""),
            sub.get("mother_name", "")
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
        row_num += 1
    
    # Footer
    if include_footer:
        row_num += 1
        ws.merge_cells(f'A{row_num}:J{row_num}')
        ws[f'A{row_num}'] = "---  End of Report  ---"
        ws[f'A{row_num}'].alignment = Alignment(horizontal='center')
        ws[f'A{row_num}'].font = Font(italic=True, color="666666")
    
    # Adjust column widths
    widths = [6, 25, 15, 10, 8, 12, 10, 12, 20, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"students_{status}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/events/{event_id}/participants")
async def export_event_participants(
    event_id: int,
    current_user: User = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_postgres_session)
):
    """Export participants for a specific event as Excel with attendance (Admin only)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")
    
    from app.models.sql_models import EventAttendance
    from datetime import timedelta
    
    # Get event
    result = await db.execute(select(Event).where(Event.id == event_id))
    event = result.scalar_one_or_none()
    if not event:
        raise HTTPException(404, "Event not found")
    
    # Get participants
    collection = get_participation_collection()
    cursor = collection.find({"event_id": event_id, "status": "selected"})
    participants = await cursor.to_list(length=500)
    
    # Get event dates
    event_dates = []
    current_date = event.start_date
    while current_date <= event.end_date:
        event_dates.append(current_date)
        current_date += timedelta(days=1)
    
    # Get attendance records
    attendance_result = await db.execute(
        select(EventAttendance).where(EventAttendance.event_id == event_id)
    )
    
    # helper for safe date string
    def to_date_str(d):
        if hasattr(d, 'strftime'):
            return d.strftime('%Y-%m-%d')
        return str(d).split(' ')[0]  # Fallback for strings or odd objects

    # Use explicit date string for robust key matching
    attendance_records = {
        (r.usn, to_date_str(r.attendance_date)): r.status 
        for r in attendance_result.scalars().all()
    }
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1a1a2e")
    present_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    absent_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells(f'A1:F1')
    ws['A1'] = f"Event: {event.name}"
    ws['A1'].font = title_font
    
    ws.merge_cells(f'A2:F2')
    ws['A2'] = f"Location: {event.location} | Date: {event.start_date} - {event.end_date}"
    
    # Headers
    headers = ["#", "USN", "Student Name", "Selected On"]
    # Add date columns for attendance
    for d in event_dates:
        headers.append(d.strftime("%d/%m") if hasattr(d, 'strftime') else str(d))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Data
    for i, p in enumerate(participants, 1):
        row = 4 + i
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=p["usn"]).border = thin_border
        ws.cell(row=row, column=3, value=p["student_name"]).border = thin_border
        ws.cell(row=row, column=4, value=str(p.get("processed_at", ""))).border = thin_border
        
        # Attendance columns
        for col_offset, att_date in enumerate(event_dates):
            # Use explicit string conversion for lookup
            date_key = to_date_str(att_date)
            status = attendance_records.get((p["usn"], date_key))
            
            cell = ws.cell(row=row, column=5 + col_offset)
            cell.border = thin_border
            
            if status and status.lower() == "present":
                cell.value = "P"
                cell.fill = present_fill
            elif status and status.lower() == "absent":
                cell.value = "Absent"
                cell.fill = absent_fill
            else:
                cell.value = "-"
    
    # Adjust widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18
    for i in range(len(event_dates)):
        ws.column_dimensions[chr(69 + i)].width = 8  # E, F, G...
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"event_{event_id}_participants_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

